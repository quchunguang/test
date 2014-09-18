#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""Import xlsx files to mysql DB, vice versa.
Filename should named like XXXyyyymmdd.xlsx (数据表20140701.xlsx for example).
"""

__author__ = 'Kevin Qu <quchunguang@gmail.com>'
__version__ = '0.4'
__nonsense__ = '6dffbe62-51c3-4aec-9189-4b46e0e63fdc'

import argparse
import os
import os.path
import re
import datetime
import sys
import MySQLdb
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter


def getws(filename):
    wb = load_workbook(filename=filename, use_iterators=True)
    ws = wb.get_sheet_by_name(name="客户商品销售明细表".decode('utf8'))
    return ws


def print_column_type(ws):
    for c in xrange(1, ws.get_highest_column()):
        name = ws.cell(row=1, column=c).value
        val = ws.cell(row=2, column=c).value
        print c, name.encode('utf8'), type(val)


def getdate(filename):
    pattern = re.compile(ur'20\d\d\d\d\d\d', re.UNICODE)
    match = pattern.search(filename)
    if match == None:
        print "[Error ] File format error: ", filename
        return None

    return datetime.datetime.strptime(match.group(), "%Y%m%d").date()


def convert_utf8(value):
    if type(value) == unicode:
        return value.encode('utf8')
    else:
        return value


def fileindb(conn, curs, filename):
    """Check if filename already inserted."""
    date = getdate(filename)
    if not date:
        return 0
    curs.execute(
        """select count(id) from imports where date=%s""", date)
    return curs.fetchall()[0][0]


def import_file(conn, curs, filename):
    """Import data to sales from target file."""
    date = getdate(filename)

    if fileindb(conn, curs, filename):
        print '[IMPORT] Already imported: ', filename
        return

    print "[IMPORT]", filename, "wait ...",

    # read data from xmls file
    ws = getws(filename)
    it = ws.iter_rows()
    it.next()

    sql_insert_sales_detail = \
        """insert into sales_detail
        (tabledate,qdmc,sdlx,khbm,sdlxbm,khmc,sptm,xsdb,plmc,ppmc,mdbm,
        mdmc,jybbm,spbm,spmc,yz,xs,ps,zs,hsxse,bhsxse,xscb,ml,mll,xsdbbm,bz)
        values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
        %s,%s,%s,%s,%s,%s)"""

    sql_insert_import = \
        """insert into imports (date,filename,importtime) values(%s,%s,%s)"""

    # insert data to sales_detail
    rows = 0
    for r in it:

        # if khmc is None, not insert!
        if r[4].value == None:
            continue

        cols = [1, 2, 3, 4, 5, 8, 9, 10, 11, 12, 13, 14, 15,
                16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, ]
        record = [convert_utf8(r[c - 1].value) for c in cols]
        record = [date] + record + [None]
        curs.execute(sql_insert_sales_detail, record)
        rows += 1

    # insert log to import
    importrecord = [date, filename, datetime.datetime.now()]
    curs.execute(sql_insert_import, importrecord)
    conn.commit()

    print rows, "record(s) inserted."


def delete_file(conn, curs, filename):
    """Delete data from db ever imported from target file."""
    if not fileindb(conn, curs, filename):
        print "[DELETE] Not need to delete: ", filename
        return
    print "[DELETE] ", filename
    tabledate = getdate(filename)
    curs.execute("""delete from sales_detail where tabledate=%s""", tabledate)
    curs.execute("""delete from imports where filename=%s""", filename)
    conn.commit()


def list_imported_file(conn, curs):
    """List all files have imported to db."""
    print "=DATE=\t\t=FILENAME=\t\t=IMPORT TIME="
    curs.execute("""select date, filename, importtime from imports""")
    for i in curs.fetchall():
        print "%s\t%s\t%s" % (i[0], i[1], i[2])


def export_folder(conn, curs, path):
    """Export data to multi xlsx files divide by `khmc`, months as sheets."""

    # contents in echo sheet of one file by following date in `tabledate`
    months = ['2013-07-01', '2013-08-01', '2013-09-01', '2013-10-01',
              '2013-11-01', '2013-12-01', '2014-01-01', '2014-02-01',
              '2014-03-01', '2014-04-01', '2014-05-01', '2014-06-01', ]

    # (column_id, column_name, column_type)
    col_info = (
        (1, '渠道名称', 's'),
        (2, '商店类型', 's'),
        (3, '客户编码', 'n'),
        (4, '商店类型编码', 's'),
        (5, '客户名称', 's'),
        (6, '商品条码', 's'),
        (7, '销售代表', 's'),
        (8, '品类名称', 's'),
        (9, '品牌名称', 's'),
        (10, '门店编码', 'n'),
        (11, '门店名称', 's'),
        (12, '经营部编码', 'n'),
        (13, '商品编码', 's'),
        (14, '商品名称', 's'),
        (15, '因子', 'n'),
        (16, '箱数', 'n'),
        (17, '片数', 'n'),
        (18, '总数', 'n'),
        (19, '含税销售额', 'n'),
        (20, '不含税销售额', 'n'),
        (21, '销售成本', 'n'),
        (22, '毛利', 'n'),
        (23, '毛利率', 'n'),
        (24, '销售代表编码', 'n'), )

    sql_khmc = """select distinct khmc from sales_detail"""
    sql_sheet = """select qdmc,sdlx,khbm,sdlxbm,khmc,sptm,xsdb,plmc,ppmc,mdbm,
mdmc,jybbm,spbm,spmc,yz,xs,ps,zs,hsxse,bhsxse,xscb,ml,mll,xsdbbm
from sales_detail where khmc=%s and tabledate=%s"""

    curs.execute(sql_khmc)
    ret = curs.fetchall()

    # each file by khml
    for khmc in ret:

        # create a new file
        dest_filename = khmc[0] + '.xlsx'
        dest_filename = os.path.join(path, dest_filename)
        if sys.platform == 'win32' or sys.platform == 'cygwin':  # Windows
            dest_filename = dest_filename.decode('utf8').encode('gbk')

        # check and ignore exists files
        if os.path.isfile(dest_filename):
            print "[EXPORT] Already exists", dest_filename
            continue

        print "[EXPORT]", dest_filename

        # create a new book
        wb = Workbook()
        ew = ExcelWriter(workbook=wb)

        # each sheet of table
        for i, month in enumerate(months):

            curs.execute(sql_sheet, (khmc[0], month))
            ret2 = curs.fetchall()

            ws = wb.create_sheet(i, month[0:-3])

            # create header
            for c, n, t in col_info:
                putvalue(ws, 1, c, n, 's')

            # fill data
            for i, r in enumerate(ret2):
                for c, n, t in col_info:
                    putvalue(ws, i + 2, c,  r[c - 1], t)

        wb.save(dest_filename)


def putvalue(ws, row, column, value, data_type):
    """Put value into a cell of xlsx sheet, given None works as well."""
    if data_type == 's':
        if value == None:
            v = u""
        else:
            v = value.decode('utf8')
    if data_type == 'n':
        if value == None:
            v = 0
        else:
            v = value
    ws.cell(row=row, column=column).set_explicit_value(
        value=v, data_type=data_type)


def import_folder(conn, curs, path):
    """Import data to db from all files in target folder."""
    file_list = []
    for f in os.listdir(path):
        root, ext = os.path.splitext(f)
        if ext == '.xlsx' and getdate(f) != None:
            file_list.append(f)

    file_list.sort()

    for f in file_list:
        import_file(conn, curs, f)


def opendb():
    """Create db connection."""
    conn = MySQLdb.connect(host='127.0.0.1',
                           user='root',
                           passwd='passwd',
                           db='sales')
    curs = conn.cursor()
    curs.execute("SET NAMES 'utf8'")
    conn.commit()
    return conn, curs


def closedb(conn, curs):
    """Close db connection."""
    curs.close()
    conn.close()


def parse_args():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(description=sys.modules[__name__].__doc__)

    parser.add_argument('-l', '--list',
                        action='store_true', default=False,
                        help=list_imported_file.__doc__)
    parser.add_argument('-d', '--delete-file', help=delete_file.__doc__)
    parser.add_argument('-e', '--export-folder', help=export_folder.__doc__)
    parser.add_argument('--import-file', help=import_file.__doc__)
    parser.add_argument('-i', '--import-folder', help=import_folder.__doc__)

    return parser.parse_args()


def main():
    conn, curs = opendb()

    args = parse_args()
    if args.list:
        list_imported_file(conn, curs)
    elif args.delete_file:
        delete_file(conn, curs, args.delete_file)
    elif args.export_folder:
        export_folder(conn, curs, args.export_folder)
    elif args.import_file:
        import_file(conn, curs, args.import_file)
    elif args.import_folder:
        import_folder(conn, curs, args.import_folder)

    closedb(conn, curs)


if __name__ == "__main__":
    main()
