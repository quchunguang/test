#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""Demo of read and write excel .xlsx file using openpyxl.
"""
__author__ = 'Kevin Qu <quchunguang@gmail.com>'
__version__ = '0.1'
__nonsense__ = '5d6906e5-0189-489d-aada-3fa75a74e041'

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

filename = r'/mnt/document/empty_book.xlsx'


def putvalue(ws, row, column, value, data_type):
    """Put a value to target cell.

    This will convert utf-8 string to unicode, and keep number as is.
    None string will convert to u"",
    None number will convert to 0.
    """
    if data_type == 's':
        if value == None:
            v = u""
        else:
            v = value.decode('utf8')
    if data_type == 'n':
        if value == None:
            v = 0
        else:
            v = value
    ws.cell(row=row, column=column).set_explicit_value(
        value=v, data_type=data_type)


def create():
    """Create new xlsx file and write something in it."""
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    for col_idx in range(1, 40):
        col = get_column_letter(col_idx)
        for row in range(1, 600):
            ws.cell('%s%s' % (col, row)).value = '%s%s' % (col, row)
    ws = wb.create_sheet()
    ws.title = 'Pi'
    ws['F5'] = 3.14
    wb.save(filename=filename)


def modify():
    """Open exist file and write something in it."""
    wb = load_workbook(filename=filename)
    ws = wb.get_sheet_by_name("data".decode('utf8'))

    # create header
    ws.cell(row=1, column=1).set_explicit_value(
        value=u'Id', data_type='s')
    ws.cell(row=1, column=2).set_explicit_value(
        value=u'Name', data_type='s')
    ws.cell(row=1, column=3).set_explicit_value(
        value=u'D1', data_type='s')
    ws.cell(row=1, column=4).set_explicit_value(
        value=u'D2', data_type='s')
    ws.cell(row=1, column=5).set_explicit_value(
        value=u'D3', data_type='s')

    # fill data
    for r in range(2, 12):
        putvalue(ws, r, 1, r - 1, 'n')
        putvalue(ws, r, 2, "品牌名称" + str(r - 1), 's')
        putvalue(ws, r, 3, (r - 1) * 3, 'n')
        putvalue(ws, r, 4, (r - 1) * 2, 'n')
        putvalue(ws, r, 5, (r - 1) * 4, 'n')
    wb.save(filename=filename)


def convert_utf8(value):
    """Convert a unicode value to utf8, otherwise, do nothing."""
    if type(value) == unicode:
        return value.encode('utf8')
    else:
        return value


def readiter():
    """Read exist file by iterator (read only)"""
    wb = load_workbook(filename=filename, use_iterators=True)
    ws = wb.get_sheet_by_name(name="data".decode('utf8'))
    it = ws.iter_rows()
    it.next()             # ignore header of table

    for row in it:
        for col in row:
            print convert_utf8(col.value),
        print 'B->', convert_utf8(row[1].value)  # row[1] means col 2


def main():
    # create()
    readiter()
    # modify()

if __name__ == "__main__":
    main()
