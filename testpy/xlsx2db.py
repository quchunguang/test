#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""Read XXX20140101.xlsx ... files to mysql db, and then write out.
"""
__author__ = 'Kevin Qu <quchunguang@gmail.com>'
__version__ = '0.1'
__nonsense__ = '6dffbe62-51c3-4aec-9189-4b46e0e63fdc'
debug = True

import os
import os.path
import re
import datetime
import MySQLdb
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter


def getws(filename):
    wb = load_workbook(filename=filename)
    ws = wb.get_sheet_by_name("客户商品销售明细表".decode('utf8'))
    return ws


def print_column_type(ws):
    for c in xrange(1, ws.get_highest_column()):
        name = ws.cell(row=1, column=c).value
        val = ws.cell(row=2, column=c).value
        print c, name.encode('utf8'), type(val)


def getdate(filename):
    pattern = re.compile(ur'\d\d\d\d\d\d\d\d', re.UNICODE)
    match = pattern.search(filename)
    if match == None:
        print "Can not get date from filename: ", filename
        return None

    return datetime.datetime.strptime(match.group(), "%Y%m%d").date()


def convert_utf8(value):
    if type(value) == unicode:
        return value.encode('utf8')
    else:
        return value


def fileindb(conn, curs, filename):
    """Check if filename already inserted"""
    date = getdate(filename)
    curs.execute(
        """select count(id) from imports where date=%s""", date)
    return curs.fetchall()[0][0]


def import_file(conn, curs, filename):
    """Import data to sales in file"""
    date = getdate(filename)

    if fileindb(conn, curs, filename):
        print '[IMPORT] Already imported: ', filename
        return

    print "[IMPORT]", filename
    # read data from xmls file
    ws = getws(filename)

    # insert data
    records = []
    for r in xrange(2, ws.get_highest_row()):
        # if khmc is None, not insert!
        if ws.cell(row=r, column=5).value == None:
            continue
        cols = [1, 2, 3, 4, 5, 8, 9, 10, 11, 12, 13, 14, 15,
                16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, ]
        record = [convert_utf8(ws.cell(row=r, column=c).value) for c in cols]
        records.append([date] + record + [None])

    curs.executemany(
        """insert into sales_detail
(tabledate,qdmc,sdlx,khbm,sdlxbm,khmc,sptm,xsdb,plmc,ppmc,mdbm,
mdmc,jybbm,spbm,spmc,yz,xs,ps,zs,hsxse,bhsxse,xscb,ml,mll,xsdbbm,bz)
values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
%s,%s,%s,%s,%s,%s)""", records)
    # conn.commit()

    # insert import log
    importrecord = [date, filename, datetime.datetime.now()]
    curs.execute(
        """insert into imports (date,filename,importtime) values(%s,%s,%s)""",
        importrecord)
    conn.commit()


def delete_by_file(conn, curs, filename):
    """Delete data from db ever imported from target file"""
    if fileindb(conn, curs, filename) == 0:
        print "[DELETE] Not need to delete: ", filename
        return
    print "[DELETE] ", filename
    tabledate = getdate(filename)
    curs.execute("""delete from sales_detail where tabledate=%s""", tabledate)
    curs.execute("""delete from imports where filename=%s""", filename)
    conn.commit()


def list_imported_file(conn, curs):
    """List all files have imported to db"""
    print "[LIST  ] Start list imported files"
    print "=DATE=\t\t=FILENAME=\t\t=IMPORT TIME="
    curs.execute("""select date, filename, importtime from imports""")
    for i in curs.fetchall():
        print "%s\t%s\t%s" % (i[0], i[1], i[2])


def export_by_khmc(conn, curs):
    """Export data to multi xlsx files divide by `khmc`"""
    curs.execute("""select distinct khmc from sales_detail""")
    ret = curs.fetchall()
    for khmc in ret:
        curs.execute("""select qdmc,sdlx,khbm,sdlxbm,khmc,sptm,xsdb,plmc,ppmc,mdbm,
mdmc,jybbm,spbm,spmc,yz,xs,ps,zs,hsxse,bhsxse,xscb,ml,mll,xsdbbm
from sales_detail where khmc=%s""", khmc[0])
        ret2 = curs.fetchall()

        # create a new file
        dest_filename = khmc[0] + '.xlsx'
        print "[EXPORT] ", dest_filename

        wb = Workbook()
        ew = ExcelWriter(workbook=wb)
        ws = wb.worksheets[0]
        ws.title = u"客户商品销售明细表"

        # create header
        ws.cell(row=1, column=1).set_explicit_value(
            value=u'渠道名称', data_type='s')
        ws.cell(row=1, column=2).set_explicit_value(
            value=u'商店类型', data_type='s')
        ws.cell(row=1, column=3).set_explicit_value(
            value=u'客户编码', data_type='s')
        ws.cell(row=1, column=4).set_explicit_value(
            value=u'商店类型编码', data_type='s')
        ws.cell(row=1, column=5).set_explicit_value(
            value=u'客户名称', data_type='s')
        ws.cell(row=1, column=6).set_explicit_value(
            value=u'商品条码', data_type='s')
        ws.cell(row=1, column=7).set_explicit_value(
            value=u'销售代表', data_type='s')
        ws.cell(row=1, column=8).set_explicit_value(
            value=u'品类名称', data_type='s')
        ws.cell(row=1, column=9).set_explicit_value(
            value=u'品牌名称', data_type='s')
        ws.cell(row=1, column=10).set_explicit_value(
            value=u'门店编码', data_type='s')
        ws.cell(row=1, column=11).set_explicit_value(
            value=u'门店名称', data_type='s')
        ws.cell(row=1, column=12).set_explicit_value(
            value=u'经营部编码', data_type='s')
        ws.cell(row=1, column=13).set_explicit_value(
            value=u'商品编码', data_type='s')
        ws.cell(row=1, column=14).set_explicit_value(
            value=u'商品名称', data_type='s')
        ws.cell(row=1, column=15).set_explicit_value(
            value=u'因子', data_type='s')
        ws.cell(row=1, column=16).set_explicit_value(
            value=u'箱数', data_type='s')
        ws.cell(row=1, column=17).set_explicit_value(
            value=u'片数', data_type='s')
        ws.cell(row=1, column=18).set_explicit_value(
            value=u'总数', data_type='s')
        ws.cell(row=1, column=19).set_explicit_value(
            value=u'含税销售额', data_type='s')
        ws.cell(row=1, column=20).set_explicit_value(
            value=u'不含税销售额', data_type='s')
        ws.cell(row=1, column=21).set_explicit_value(
            value=u'销售成本', data_type='s')
        ws.cell(row=1, column=22).set_explicit_value(
            value=u'毛利', data_type='s')
        ws.cell(row=1, column=23).set_explicit_value(
            value=u'毛利率', data_type='s')
        ws.cell(row=1, column=24).set_explicit_value(
            value=u'销售代表编码', data_type='s')

        for i, r in enumerate(ret2):
            putvalue(ws, i + 2, 1, r[0], 's')
            putvalue(ws, i + 2, 2, r[1], 's')
            putvalue(ws, i + 2, 3, r[2], 'n')
            putvalue(ws, i + 2, 4, r[3], 's')
            putvalue(ws, i + 2, 5, r[4], 's')
            putvalue(ws, i + 2, 6, r[5], 's')
            putvalue(ws, i + 2, 7, r[6], 's')
            putvalue(ws, i + 2, 8, r[7], 's')
            putvalue(ws, i + 2, 9, r[8], 's')
            putvalue(ws, i + 2, 10, r[9], 'n')
            putvalue(ws, i + 2, 11, r[10], 's')
            putvalue(ws, i + 2, 12, r[11], 'n')
            putvalue(ws, i + 2, 13, r[12], 's')
            putvalue(ws, i + 2, 14, r[13], 's')
            putvalue(ws, i + 2, 15, r[14], 'n')
            putvalue(ws, i + 2, 16, r[15], 'n')
            putvalue(ws, i + 2, 17, r[16], 'n')
            putvalue(ws, i + 2, 18, r[17], 'n')
            putvalue(ws, i + 2, 19, r[18], 'n')
            putvalue(ws, i + 2, 20, r[19], 'n')
            putvalue(ws, i + 2, 21, r[20], 'n')
            putvalue(ws, i + 2, 22, r[21], 'n')
            putvalue(ws, i + 2, 23, r[22], 'n')
            putvalue(ws, i + 2, 24, r[23], 'n')
        wb.save(dest_filename)


def putvalue(ws, row, column, value, data_type):
    if data_type == 's':
        if value == None:
            v = u""
        else:
            v = value.decode('utf8')
    if data_type == 'n':
        if value == None:
            v = 0
        else:
            v = value
    ws.cell(row=row, column=column).set_explicit_value(
        value=v, data_type=data_type)


def import_folder(conn, curs, path):
    """Import data to db from all files in target folder"""
    for f in os.listdir(path):
        root, ext = os.path.splitext(f)
        if ext == '.xlsx' and getdate(f) != None:
            import_file(conn, curs, f)


def opendb():
    """Create db connection"""
    conn = MySQLdb.connect(host='127.0.0.1',
                           user='root',
                           passwd='passwd',
                           db='sales')
    curs = conn.cursor()
    curs.execute("SET NAMES 'utf8'")
    conn.commit()
    return conn, curs


def closedb(conn, curs):
    """Close db connection"""
    curs.close()
    conn.close()


def main():
    conn, curs = opendb()

    path = '/home/qcg/share/test/testpy/'
    import_folder(conn, curs, path)
    # export_by_khmc(conn, curs)

    # filename = '数据分析表20140912.xlsx'
    # delete_by_file(conn, curs, filename)
    list_imported_file(conn, curs)

    closedb(conn, curs)


if __name__ == "__main__":
    main()
